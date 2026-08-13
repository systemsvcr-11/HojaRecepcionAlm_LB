import io
import os
from copy import copy
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, PatternFill, Side
from openpyxl.utils import get_column_letter

# ReportLab para PDF
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Table, TableStyle

MAPEO_AREAS = {
    "La Benita": {
        "COCINA BENITA": "COCINA",
        "Gisela Benitez": "ATENCION",
        "Ariana Maquera": "BARRA",
        "Angie Apaza": "CAJA",
    },
    "La Guardiana": {
        "COCINA GUARDIANA": "COCINA",
        "BARRA GUARDIANA": "BARRA",
        "CAJA GUARDIANA": "CAJA",
    },
    "Centro de Producción": {
        "PRODUCCION": "PRODUCCION",
        "ALMACEN": "ALMACEN",
    },
}

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
RUTA_BD_GENERAL = os.path.join(
    BASE_DIR, "bd", "BD MERGE UNIDADES GENERAL 5.xlsx"
)
RUTA_BD_CORRELACIONES = os.path.join(
    BASE_DIR, "bd", "MERGE CORRELACIONES 1.xlsx"
)
RUTA_BD_UBICACIONES = os.path.join(
    BASE_DIR, "bd", "UBICACIONES PAS.xlsx"
)


def procesar_archivos_excel(archivos_subidos, local_seleccionado):
    """Procesa los archivos Odoo realizando los MERGES de BD General, Correlaciones y Ubicaciones PAS."""
    lista_df = [pd.read_excel(f) for f in archivos_subidos]
    df = pd.concat(lista_df, ignore_index=True)

    columnas_ffill = [
        "ID",
        "Referencia de la orden",
        "Cliente",
        "Comprador",
        "Fecha esperada",
        "Líneas de la orden/Producto",
        "Líneas de la orden/Cantidad",
        "Líneas de la orden/Unidad",
        "Líneas de la orden/Descripción",
        "Líneas de la orden/Producto/ID",
    ]

    cols_existentes = [c for c in columnas_ffill if c in df.columns]
    df[cols_existentes] = df[cols_existentes].ffill()

    mapeo_actual = MAPEO_AREAS.get(local_seleccionado, {})

    if "Comprador" in df.columns:
        df["AREA"] = df["Comprador"].map(mapeo_actual).fillna(df["Comprador"])
        df = df.rename(columns={"Comprador": "Local"})

    df["RQ"] = (
        df["Referencia de la orden"]
        if "Referencia de la orden" in df.columns
        else ""
    )

    if "Fecha esperada" in df.columns:
        df["Fecha esperada"] = pd.to_datetime(
            df["Fecha esperada"], errors="coerce"
        ).dt.strftime("%d/%m/%Y")

    df = df.rename(
        columns={
            "Líneas de la orden/Producto": "Producto",
            "Líneas de la orden/Unidad": "Unidad",
            "Líneas de la orden/Cantidad": "Cantidad",
            "Líneas de la orden/Descripción": "Observaciones",
        }
    )

    df["Ubicacion"] = ""
    df["Cant 1"] = ""
    df["Falta"] = ""
    df["FIFO 1"] = ""
    df["FIFO 2"] = ""

    cols_deseadas = [
        "RQ",
        "AREA",
        "Fecha esperada",
        "Producto",
        "Unidad",
        "Ubicacion",
        "Cantidad",
        "Cant 1",
        "Falta",
        "FIFO 1",
        "FIFO 2",
        "Observaciones",
        "ID",
        "Líneas de la orden/Producto/ID",
    ]

    cols_finales = [c for c in cols_deseadas if c in df.columns]
    df = df[cols_finales]

    if "Cantidad" in df.columns:
        df = df[df["Cantidad"] != 0.0]

    if "Producto" in df.columns and "Observaciones" in df.columns:
        df.loc[df["Producto"] == df["Observaciones"], "Observaciones"] = ""
        df["Observaciones"] = df.apply(
            lambda row: ""
            if pd.isna(row["Observaciones"]) or pd.isna(row["Producto"])
            else str(row["Observaciones"])
            .replace(str(row["Producto"]), "")
            .replace("\n", " "),
            axis=1,
        )

    # =========================================================
    # MERGE 1: BD GENERAL
    # =========================================================
    if os.path.exists(RUTA_BD_GENERAL):
        df_bd = pd.read_excel(RUTA_BD_GENERAL)
        df_bd = df_bd.rename(columns={"ITEM NAME ODOO": "Producto"})

        df = df.merge(df_bd, on="Producto", how="left")
        df["Producto"] = df["Producto"].astype(str) + df["Observaciones"].astype(str)

        cols_m1 = [
            "RQ",
            "AREA",
            "CATEGORIA",
            "Fecha esperada",
            "Producto",
            "Unidad",
            "Ubicacion",
            "Cantidad",
            "Cant 1",
            "Falta",
            "FIFO 1",
            "FIFO 2",
            "ID",
            "Líneas de la orden/Producto/ID",
        ]
        df = df[[c for c in cols_m1 if c in df.columns]]
        if "CATEGORIA" in df.columns:
            df = df.sort_values(by=["RQ", "CATEGORIA"], kind="mergesort")

    # =========================================================
    # MERGE 2: MERGE CORRELACIONES (Conversión de Unidades)
    # =========================================================
    if os.path.exists(RUTA_BD_CORRELACIONES):
        df_corr = pd.read_excel(RUTA_BD_CORRELACIONES)
        df_corr.rename(columns={"ITEM NAME ODOO": "Producto"}, inplace=True)

        df = pd.merge(df, df_corr, on="Producto", how="left")

        if "RENDIMIENTO 2" in df.columns:
            df["Cantidad RQ"] = df["Cantidad"] * df["RENDIMIENTO 2"]
            df.drop(columns=["Unidad", "Cantidad"], inplace=True, errors="ignore")
            df.rename(
                columns={
                    "Cantidad RQ": "Cantidad",
                    "UNIDAD INTERNA 2": "Unidad",
                    "CATEGORIA_x": "CATEGORIA",
                },
                inplace=True,
            )

        cols_a_borrar = [
            "ESTADO", "PROVEEDOR", "CATEGORIA_y", "VALOR", "SUB UNIDAD",
            "SUB CANTIDAD", "CANTIDAD UNIDAD INTERNA", "STOCK TOTAL",
            "CANTIDAD 1", "UNIDAD INTERNA 1", "RENDIMIENTO 1",
            "CANTIDAD CONVERTIDA", "UNIDAD ODOO\n(UNIDAD PEDIDO)",
            "CANTIDAD 2", "RENDIMIENTO 2", "CANTIDAD CONVERTIDA 2",
            "ESTADO CONVERSION", "CANTIDAD CONVERTIDA ",
        ]
        df.drop(columns=cols_a_borrar, inplace=True, errors="ignore")

    # =========================================================
    # MERGE 3: UBICACIONES PAS
    # =========================================================
    if os.path.exists(RUTA_BD_UBICACIONES) and "CATEGORIA" in df.columns:
        df_ubic = pd.read_excel(RUTA_BD_UBICACIONES)

        # Normalizar nombres de encabezados del archivo de ubicaciones
        df_ubic.columns = df_ubic.columns.astype(str).str.strip()

        # Buscar la columna del producto/categoría en UBICACIONES PAS
        col_cat_ubic = None
        for col in df_ubic.columns:
            if "CATEGORIA" in col.upper() or "PAS" in col.upper():
                col_cat_ubic = col
                break

        col_ubic_val = None
        for col in df_ubic.columns:
            if "UBIC" in col.upper():
                col_ubic_val = col
                break

        if col_cat_ubic and col_ubic_val:
            df_ubic.rename(
                columns={
                    col_cat_ubic: "CATEGORIA_PAS",
                    col_ubic_val: "Ubicacion_PAS",
                },
                inplace=True,
            )

            # Limpieza profunda de texto: eliminar espacios y homogeneizar a mayúsculas
            df["CATEGORIA_CLEAN"] = (
                df["CATEGORIA"].astype(str).str.strip().str.upper()
            )
            df_ubic["CATEGORIA_PAS_CLEAN"] = (
                df_ubic["CATEGORIA_PAS"].astype(str).str.strip().str.upper()
            )

            # Realizar el cruce con las llaves limpias
            df = pd.merge(
                df,
                df_ubic[["CATEGORIA_PAS_CLEAN", "Ubicacion_PAS"]],
                left_on="CATEGORIA_CLEAN",
                right_on="CATEGORIA_PAS_CLEAN",
                how="left",
            )

            # Asignar el resultado a la columna Ubicacion y limpiar temporales
            if "Ubicacion_PAS" in df.columns:
                df["Ubicacion"] = df["Ubicacion_PAS"].fillna("")
                df.drop(
                    columns=[
                        "CATEGORIA_CLEAN",
                        "CATEGORIA_PAS_CLEAN",
                        "Ubicacion_PAS",
                    ],
                    inplace=True,
                    errors="ignore",
                )

    # Redondear la columna Cantidad a 1 decimal
    if "Cantidad" in df.columns:
        df["Cantidad"] = pd.to_numeric(df["Cantidad"], errors="coerce").round(1)

    cols_orden_final = [
        "RQ",
        "AREA",
        "CATEGORIA",
        "Fecha esperada",
        "Producto",
        "Unidad",
        "Ubicacion",
        "Cantidad",
        "Cant 1",
        "Falta",
        "FIFO 1",
        "FIFO 2",
        "ID",
        "Líneas de la orden/Producto/ID",
    ]
    df = df[[c for c in cols_orden_final if c in df.columns]]

    return df


def generar_excel_formateado(df):
    """Genera archivo Excel aplicando estilos."""
    output = io.BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Datos")

    output.seek(0)
    wb = load_workbook(output)
    ws = wb["Datos"]

    fill_g = PatternFill(
        start_color="D8E4BC", end_color="D8E4BC", fill_type="solid"
    )
    fill_j = PatternFill(
        start_color="FCD5B4", end_color="FCD5B4", fill_type="solid"
    )
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    headers = {cell.value: idx for idx, cell in enumerate(ws[1], start=1)}

    if "Ubicacion" in headers:
        col_u = headers["Ubicacion"]
        for r in range(2, ws.max_row + 1):
            ws.cell(row=r, column=col_u).fill = fill_g

    if "Falta" in headers:
        col_f = headers["Falta"]
        for r in range(2, ws.max_row + 1):
            ws.cell(row=r, column=col_f).fill = fill_j

    for row in ws.iter_rows(
        min_row=1, max_row=ws.max_row, max_col=ws.max_column
    ):
        for cell in row:
            cell.border = thin_border

    center_align = Alignment(horizontal="center", vertical="center")
    for col_name in ["Fecha esperada", "Unidad", "Cantidad"]:
        if col_name in headers:
            col_idx = headers[col_name]
            for row in range(2, ws.max_row + 1):
                ws.cell(row=row, column=col_idx).alignment = center_align

    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max_len + 3

    colores = [
        "D9EAD3", "D0E0E3", "FCE5CD", "EAD1DC", "FFF2CC", "D9D2E9", "CFE2F3"
    ]
    mapa_colores = {}
    color_idx = 0

    for row in range(2, ws.max_row + 1):
        valor = ws.cell(row=row, column=1).value
        if valor not in mapa_colores:
            mapa_colores[valor] = colores[color_idx % len(colores)]
            color_idx += 1

        fill = PatternFill(
            start_color=mapa_colores[valor],
            end_color=mapa_colores[valor],
            fill_type="solid",
        )
        ws.cell(row=row, column=1).fill = fill
        if ws.max_column >= 2:
            ws.cell(row=row, column=2).fill = fill

    for col_hid in ["ID", "Líneas de la orden/Producto/ID"]:
        if col_hid in headers:
            ws.column_dimensions[
                get_column_letter(headers[col_hid])
            ].hidden = True

    final_output = io.BytesIO()
    wb.save(final_output)
    final_output.seek(0)
    return final_output


def generar_pdf_reportlab(df, local_seleccionado):
    """Genera PDF idéntico a Excel pero ultra-compacto para minimizar páginas."""
    pdf_buffer = io.BytesIO()

    # 1. MÁRGENES MÍNIMOS (0.3 cm por lado y 1.1 cm superior para el título)
    margin = 0.3 * cm

    doc = SimpleDocTemplate(
        pdf_buffer,
        pagesize=landscape(A4),
        leftMargin=margin,
        rightMargin=margin,
        topMargin=1.1 * cm,  # Espacio reducido para título
        bottomMargin=0.5 * cm,
    )

    fecha_str = datetime.now().strftime("%d.%m")
    titulo_texto = f"RQ {local_seleccionado.upper()} - {fecha_str}"

    # 2. ENCABEZADO MÁS COMPACTO
    def agregar_encabezado_pagina(canvas, doc):
        canvas.saveState()
        canvas.setFont("Helvetica-Bold", 12)  # Bajamos de 16 a 12pt
        canvas.drawCentredString(
            landscape(A4)[0] / 2.0,
            landscape(A4)[1] - 0.8 * cm,
            titulo_texto,
        )
        canvas.restoreState()

    elements = []

    cols_pdf = [
        c
        for c in df.columns
        if c not in ["ID", "Líneas de la orden/Producto/ID"]
    ]
    df_pdf = df[cols_pdf]

    HEX_COLORES_RQ = [
        "#D9EAD3",
        "#D0E0E3",
        "#FCE5CD",
        "#EAD1DC",
        "#FFF2CC",
        "#D9D2E9",
        "#CFE2F3",
    ]
    mapa_colores_rq = {}
    color_idx = 0

    if "RQ" in df_pdf.columns:
        for valor in df_pdf["RQ"].unique():
            mapa_colores_rq[valor] = HEX_COLORES_RQ[
                color_idx % len(HEX_COLORES_RQ)
            ]
            color_idx += 1

    # 3. FUENTES Y LEADING (INTERLINEADO) COMPACTOS
    # Bajamos el tamaño de letra a 6pt y leading a 7pt para que el texto multilinea ocupe muy poca altura
    style_header = ParagraphStyle(
        "PDFHeader",
        fontName="Helvetica-Bold",
        fontSize=7,
        leading=8,
        alignment=1,
        textColor=colors.whitesmoke,
    )
    style_cell_left = ParagraphStyle(
        "PDFCellLeft",
        fontName="Helvetica",
        fontSize=6,
        leading=7,
        alignment=0,
    )
    style_cell_center = ParagraphStyle(
        "PDFCellCenter",
        fontName="Helvetica",
        fontSize=6,
        leading=7,
        alignment=1,
    )
    style_cell_unidad = ParagraphStyle(
        "PDFCellUnidad",
        fontName="Helvetica",
        fontSize=6,
        leading=7,
        alignment=1,
        wordWrap=None,
    )

    cols_centradas = ["Fecha esperada", "Cantidad"]
    formatted_data = [[Paragraph(str(c), style_header) for c in cols_pdf]]

    for _, row in df_pdf.iterrows():
        row_cells = []
        for col_name in cols_pdf:
            text = "" if pd.isna(row[col_name]) else str(row[col_name])

            if col_name == "Unidad":
                style = style_cell_unidad
            elif col_name in cols_centradas:
                style = style_cell_center
            else:
                style = style_cell_left

            row_cells.append(Paragraph(text, style))
        formatted_data.append(row_cells)

    # 4. PADDING (RELLENO DE CELDAS) AL MÍNIMO
    table_styles = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#343A40")),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 1.2),  # Reducido de 2.5 a 1.2
        ("BOTTOMPADDING", (0, 0), (-1, -1), 1.2),  # Reducido de 2.5 a 1.2
        ("LEFTPADDING", (0, 0), (-1, -1), 1),
        ("RIGHTPADDING", (0, 0), (-1, -1), 1),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#A0A0A0")),
    ]

    idx_ubicacion = (
        cols_pdf.index("Ubicacion") if "Ubicacion" in cols_pdf else -1
    )
    idx_falta = cols_pdf.index("Falta") if "Falta" in cols_pdf else -1

    for row_idx, row in enumerate(df_pdf.iterrows(), start=1):
        rq_val = row[1]["RQ"] if "RQ" in df_pdf.columns else ""
        hex_color = mapa_colores_rq.get(rq_val, "#FFFFFF")

        table_styles.append(
            (
                "BACKGROUND",
                (0, row_idx),
                (1, row_idx),
                colors.HexColor(hex_color),
            )
        )

        if idx_ubicacion != -1:
            table_styles.append(
                (
                    "BACKGROUND",
                    (idx_ubicacion, row_idx),
                    (idx_ubicacion, row_idx),
                    colors.HexColor("#D8E4BC"),
                )
            )

        if idx_falta != -1:
            table_styles.append(
                (
                    "BACKGROUND",
                    (idx_falta, row_idx),
                    (idx_falta, row_idx),
                    colors.HexColor("#FCD5B4"),
                )
            )

    page_width = landscape(A4)[0] - (margin * 2)

    widths_ratio = {
        "RQ": 0.06,
        "AREA": 0.06,
        "CATEGORIA": 0.08,
        "Fecha esperada": 0.07,
        "Producto": 0.18,
        "Unidad": 0.06,
        "Ubicacion": 0.08,
        "Cantidad": 0.05,
        "Cant 1": 0.05,
        "Falta": 0.05,
        "FIFO 1": 0.05,
        "FIFO 2": 0.05,
        "Observaciones": 0.16,
    }

    col_widths = [
        page_width * widths_ratio.get(col, 1 / len(cols_pdf)) for col in cols_pdf
    ]

    t = Table(formatted_data, colWidths=col_widths, repeatRows=1)
    t.setStyle(TableStyle(table_styles))
    elements.append(t)

    doc.build(
        elements,
        onFirstPage=agregar_encabezado_pagina,
        onLaterPages=agregar_encabezado_pagina,
    )

    pdf_buffer.seek(0)
    return pdf_buffer
